"""Tier-0 fillable RFQ quote form — deterministic render + hardened round-trip parse
(ADR-0004 decision 10, slice E6 / PR-B).

Purpose
-------
The RFQ lane (Lane 2) attaches a fillable `.xlsx` quote form to every outbound RFQ;
a vendor fills the price cells and the office uploads the returned file through the
ordinary estimate-upload pool. This module is BOTH halves of that round trip:

* `render_quote_form` — openpyxl-render the form: a visible **'Quote Form'** sheet
  (header block: RFQ number / job / vendor / due date; a line table `# | Part
  Number | Description | Qty | Unit | Unit Price | Extended` prefilled from the
  RFQ's requested lines, price cells BLANK with a currency number format) plus a
  **`_ITS_META`** sheet (`sheet_state='veryHidden'`) carrying the form identity as
  three defined names — `ITS_RFQ_NUMBER`, `ITS_VENDOR_KEY`, and `ITS_FORM_TOKEN`
  (= `shared.portal_hmac.rfq_form_token`, the `rfq-form:v1` MAC over
  rfq_number + vendor_key). Output is BYTE-DETERMINISTIC for fixed inputs (the
  `subcontract_docx._normalize_ooxml_clock` OOXML clock pin, mirrored below —
  §42 documented copy of a module-private helper, the po_attach_screen posture).

* `parse_quote_form` — parse a RETURNED (hostile) form: locate the defined names,
  constant-time verify the token (`portal_hmac.verify_rfq_form_token`), and read
  the line table back with the red-team #3 hardening:

    - Any NUMERIC input cell (Qty / Unit Price / Extended) whose RAW value is a
      formula cell or a string starting with `=` `+` `-` `@` TAB or CR **rejects
      that cell → the WHOLE form falls to None** (never parse a formula as a
      number; a rejected form degrades to the ordinary extraction ladder).
    - Every TEXT field carried forward (part number / description / unit) has
      formula-lead characters STRIPPED (the CSV/xlsx-injection neutralization)
      and is length-capped to the Worker's extraction contract.
    - An absent/tampered token → the parse still returns the lines but
      `verified=False` and `rfq_number`/`vendor_key` are None — an ORDINARY
      ladder upload, never an auto-bind (the token asserts identity only; every
      filled value stays untrusted data either way).

  Money parses through `po_materials.estimate_parse.to_cents` (the sibling-owned
  integer-cents money parser — imported lazily so this module, and through it the
  `estimate_poll` daemon, stays importable while the extraction-core PR is in
  flight).

Trust boundary
--------------
A returned form is HOSTILE bytes. The `estimate_poll` daemon runs the §34 screen
(`po_attach_screen.screen_attachment` — zip-bomb caps, macro payloads, the widened
external-relationship gate) BEFORE this module ever sees the file; openpyxl here
only ever opens a screen-clean container. `keep_vba=False` always; `data_only=True`
for values (a formula's cached value is never trusted — the RAW workbook detects
the formula and rejects, see above).

Failure modes
-------------
`parse_quote_form` returns None on: an unreadable/oversized workbook, a missing
'Quote Form' sheet, or ANY rejected numeric cell. It NEVER raises on hostile
input. A None means "not a parsable filled form" — the caller (the ladder head in
`estimate_poll`) degrades the document to the ordinary tiers / needs_review.

Consumers
---------
`po_materials.estimate_poll._attempt_extraction_ladder` (Tier 0, this PR) and the
RFQ generator `rfq_generate` (Lane 2 slice R2, future) which calls
`render_quote_form` per (rfq, vendor). Offline: `scripts/eval_estimate_ladder.py`.
"""
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from shared import portal_hmac

# The schema_version stamped on Tier-0 extraction payloads (Worker cap: 32 chars).
FORM_SCHEMA_VERSION = "rfq-form:v1"

# ---- Fixed sheet geometry (parse relies on it; the form is OUR artifact) ----------
SHEET_FORM = "Quote Form"
SHEET_META = "_ITS_META"

HEADER_LABEL_ROWS = {  # visible header block: (row, label) → value in column B
    "rfq_number": 3,
    "job_name": 4,
    "vendor_key": 5,
    "due_date": 6,
}
TABLE_HEADER_ROW = 8
FIRST_LINE_ROW = 9
COL_POSITION = 1     # A — "#"
COL_PART_NUMBER = 2  # B
COL_DESCRIPTION = 3  # C
COL_QTY = 4          # D
COL_UNIT = 5         # E
COL_UNIT_PRICE = 6   # F — vendor-filled, currency format
COL_EXTENDED = 7     # G — vendor-filled or blank, currency format
TABLE_HEADERS = ("#", "Part Number", "Description", "Qty", "Unit", "Unit Price", "Extended")

CURRENCY_FORMAT = '"$"#,##0.00'

# Worker extraction-contract caps (safety_portal/worker/po_estimates.ts) — text
# carried forward is truncated to fit, never rejected for length.
MAX_LINES = 500        # Worker MAX_LINES
MAX_DESCRIPTION = 512  # Worker MAX_LINE_TEXT
MAX_SHORT = 64         # Worker MAX_SHORT (part numbers)
MAX_UNIT = 32
MAX_QTY = 1_000_000_000
MAX_CENTS = 10_000_000_000  # defensive local cap; Worker isCents bounds similarly

# Parse ceiling on the workbook itself (defense-in-depth behind the §34 screen).
MAX_FORM_BYTES = 10_000_000

# Formula-lead characters (red-team #3): a NUMERIC cell whose raw string starts
# with one of these rejects the whole form; a TEXT field has them stripped.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")

# _ITS_META cell coordinates the defined names point at (and the direct fallback).
_META_CELLS = {"ITS_RFQ_NUMBER": "B1", "ITS_VENDOR_KEY": "B2", "ITS_FORM_TOKEN": "B3"}

# Deterministic-output clock pin: a FIXED stamp (not wall-clock) so identical
# inputs render identical bytes (§47 idempotent filing / golden tests).
_FORM_CLOCK_STAMP = datetime(2026, 1, 1)

_CORE_DT_RE = re.compile(rb"(<dcterms:(created|modified)[^>]*>)[^<]+(</dcterms:\2>)")


@dataclass(frozen=True)
class FormParseResult:
    """One parsed returned quote form.

    `verified` is True only when the `rfq-form:v1` token verified — the ONLY state
    in which `rfq_number`/`vendor_key` are populated (the auto-bind identity).
    `lines` are Worker-extraction-contract line dicts (position / part_number /
    description / qty / unit / unit_cost_cents / extended_cents / math_ok 0|1 /
    section / line_note); `math_ok` is 1 iff every PRICED line's arithmetic
    checked; `subtotal_cents` sums the priced lines' extended cents.
    """

    verified: bool
    rfq_number: str | None
    vendor_key: str | None
    lines: list[dict[str, Any]]
    math_ok: int
    subtotal_cents: int | None


# ---- Determinism helper (documented copy, §42) -------------------------------------


def _normalize_ooxml_clock(data: bytes, stamp: datetime) -> bytes:
    """Rebuild the OOXML ZIP with BOTH wall-clock sources pinned to `stamp` — a
    documented copy of `subcontracts.subcontract_docx._normalize_ooxml_clock`
    (module-private there; the po_attach_screen copy posture, §42):

      1. every member's ZIP local-header `date_time` (openpyxl stamps these from
         wall-clock at save); and
      2. `docProps/core.xml`'s `<dcterms:created>`/`<dcterms:modified>` CONTENT
         (openpyxl overwrites `modified` with `now()` regardless of properties).

    Member order + all other content preserved; deterministic for fixed input."""
    dt = (min(2107, max(1980, stamp.year)), stamp.month, stamp.day, 0, 0, 0)
    iso = f"{stamp.year:04d}-{stamp.month:02d}-{stamp.day:02d}T00:00:00Z".encode()
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data), "r") as src, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            content = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                content = _CORE_DT_RE.sub(rb"\g<1>" + iso + rb"\g<3>", content)
            zi = zipfile.ZipInfo(item.filename, date_time=dt)
            zi.compress_type = item.compress_type
            zi.external_attr = item.external_attr
            zi.internal_attr = item.internal_attr
            zi.create_system = item.create_system
            dst.writestr(zi, content)
    return out.getvalue()


# ---- Render -------------------------------------------------------------------------


def render_quote_form(
    rfq_number: str,
    vendor_key: str,
    job_name: str,
    lines: list[dict],
    *,
    secret: bytes,
    due_date: str | None = None,
) -> bytes:
    """Render one fillable quote form for `(rfq_number, vendor_key)`.

    `lines` are the RFQ's requested lines — dicts with `part_number` /
    `description` / `qty` / `unit` (any may be absent). The vendor fills the
    Unit Price (and optionally Extended) cells; everything else arrives
    prefilled. Deterministic bytes for fixed inputs (see module docstring).
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET_FORM

    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Request for Quote — Quote Form").font = title_font
    header_values = {
        "rfq_number": ("RFQ Number:", rfq_number),
        "job_name": ("Job:", job_name),
        "vendor_key": ("Vendor:", vendor_key),
        "due_date": ("Quote Due:", due_date or ""),
    }
    for key, (label, value) in header_values.items():
        row = HEADER_LABEL_ROWS[key]
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)

    for col, header in enumerate(TABLE_HEADERS, start=1):
        cell = ws.cell(row=TABLE_HEADER_ROW, column=col, value=header)
        cell.font = label_font
        cell.alignment = Alignment(horizontal="center")

    for index, line in enumerate(lines[:MAX_LINES]):
        row = FIRST_LINE_ROW + index
        ws.cell(row=row, column=COL_POSITION, value=index + 1)
        ws.cell(row=row, column=COL_PART_NUMBER, value=str(line.get("part_number") or ""))
        ws.cell(row=row, column=COL_DESCRIPTION, value=str(line.get("description") or ""))
        qty = line.get("qty")
        if qty is not None:
            ws.cell(row=row, column=COL_QTY, value=qty)
        ws.cell(row=row, column=COL_UNIT, value=str(line.get("unit") or ""))
        # Price cells: BLANK, currency-formatted — the vendor's input surface.
        ws.cell(row=row, column=COL_UNIT_PRICE).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=COL_EXTENDED).number_format = CURRENCY_FORMAT

    widths = {1: 6, 2: 18, 3: 48, 4: 10, 5: 8, 6: 14, 7: 14}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Hidden identity sheet + defined names (the rfq-form:v1 token).
    meta = wb.create_sheet(SHEET_META)
    token = portal_hmac.rfq_form_token(secret, rfq_number, vendor_key)
    meta["A1"], meta["B1"] = "ITS_RFQ_NUMBER", rfq_number
    meta["A2"], meta["B2"] = "ITS_VENDOR_KEY", vendor_key
    meta["A3"], meta["B3"] = "ITS_FORM_TOKEN", token
    meta.sheet_state = "veryHidden"
    for name, cell in _META_CELLS.items():
        col_letter = cell[0]
        row_no = cell[1:]
        wb.defined_names.add(DefinedName(
            name, attr_text=f"'{SHEET_META}'!${col_letter}${row_no}"
        ))

    buf = io.BytesIO()
    wb.save(buf)
    return _normalize_ooxml_clock(buf.getvalue(), _FORM_CLOCK_STAMP)


# ---- Parse --------------------------------------------------------------------------


def _neutralize_text(value: Any, cap: int) -> str:
    """Carry a TEXT field forward with formula-lead chars stripped (red-team #3)
    and the Worker length cap applied. None-safe."""
    if value is None:
        return ""
    text = str(value)
    while text and text[0] in _FORMULA_LEAD:
        text = text[1:]
    return text.strip()[:cap]


def _numeric_cell_rejected(raw_cell: Any) -> bool:
    """True when a NUMERIC input cell must reject the whole form: the raw cell is
    a formula, or its raw value is a string starting with a formula-lead char.
    (data_only=True would silently hand us a formula's CACHED value — or None —
    so the RAW workbook is the authority on what the cell actually contains.)"""
    if getattr(raw_cell, "data_type", None) == "f":
        return True
    value = raw_cell.value
    return isinstance(value, str) and value.startswith(_FORMULA_LEAD)


def _cell_to_qty(value: Any) -> float | None:
    """Qty cell → bounded non-negative float, or None. The formula/lead-char
    rejection already ran; a residual string here is 'plain text that is not a
    number' and degrades to None."""
    if value is None:
        return None
    try:
        qty = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    if not (0 <= qty <= MAX_QTY):
        return None
    return qty


def _round_half_up(value: float) -> int:
    """Round half away from zero for non-negative money math (the po_generate
    _js_round posture — never banker's rounding on cents)."""
    return int(value + 0.5)


def _read_meta(wb_val: Any) -> dict[str, str]:
    """Resolve the three identity values via the defined names (fallback: the
    fixed `_ITS_META` cells). Missing/unreadable entries are simply absent."""
    out: dict[str, str] = {}
    for name, fallback_cell in _META_CELLS.items():
        value = None
        try:
            dn = wb_val.defined_names.get(name)
        except (KeyError, TypeError):
            dn = None
        if dn is not None:
            try:
                for sheet_title, coord in dn.destinations:
                    if sheet_title in wb_val.sheetnames:
                        value = wb_val[sheet_title][coord.replace("$", "")].value
                        break
            except (KeyError, ValueError, TypeError, AttributeError):
                value = None
        if value is None and SHEET_META in wb_val.sheetnames:
            try:
                value = wb_val[SHEET_META][fallback_cell].value
            except (KeyError, ValueError, TypeError):
                value = None
        if isinstance(value, str) and value:
            out[name] = value
    return out


def parse_quote_form(data: bytes, *, secret: bytes) -> FormParseResult | None:
    """Parse a RETURNED quote form (hostile bytes; §34-screened by the caller).

    Returns None when the workbook is unreadable, the 'Quote Form' sheet is
    absent, or ANY numeric input cell was rejected (formula / formula-lead
    string — the whole-form rejection, red-team #3). Otherwise returns a
    `FormParseResult`; `verified=False` (token absent/tampered) keeps the lines
    but carries no RFQ identity. Never raises on hostile input.
    """
    # Money parser is sibling-owned (estimate_parse, the extraction-core PR) —
    # lazy import keeps this module importable while that PR is in flight.
    from po_materials import estimate_parse  # noqa: PLC0415 — deliberate lazy sibling import

    if not data or len(data) > MAX_FORM_BYTES:
        return None
    try:
        wb_val = load_workbook(io.BytesIO(data), data_only=True, keep_vba=False)
        wb_raw = load_workbook(io.BytesIO(data), data_only=False, keep_vba=False)
    except Exception:  # noqa: BLE001 — hostile container; None is the degrade signal
        return None
    try:
        if SHEET_FORM not in wb_val.sheetnames or SHEET_FORM not in wb_raw.sheetnames:
            return None
        ws_val = wb_val[SHEET_FORM]
        ws_raw = wb_raw[SHEET_FORM]

        lines: list[dict[str, Any]] = []
        all_math_ok = True
        subtotal_cents = 0
        any_priced = False
        for offset in range(MAX_LINES):
            row = FIRST_LINE_ROW + offset
            part_raw = ws_raw.cell(row=row, column=COL_PART_NUMBER)
            desc_raw = ws_raw.cell(row=row, column=COL_DESCRIPTION)
            unit_raw = ws_raw.cell(row=row, column=COL_UNIT)
            qty_raw = ws_raw.cell(row=row, column=COL_QTY)
            price_raw = ws_raw.cell(row=row, column=COL_UNIT_PRICE)
            ext_raw = ws_raw.cell(row=row, column=COL_EXTENDED)

            row_empty = all(
                c.value is None
                for c in (part_raw, desc_raw, unit_raw, qty_raw, price_raw, ext_raw)
            )
            if row_empty:
                break

            # Red-team #3 — numeric cells: formula or formula-lead string rejects
            # the WHOLE form (never parse a formula as a number).
            for cell in (qty_raw, price_raw, ext_raw):
                if _numeric_cell_rejected(cell):
                    return None

            description = _neutralize_text(desc_raw.value, MAX_DESCRIPTION)
            part_number = _neutralize_text(part_raw.value, MAX_SHORT)
            unit = _neutralize_text(unit_raw.value, MAX_UNIT)
            if not description:
                # A line with no description cannot enter the extraction contract
                # (Worker requires one); fold the part number in if that's all
                # the vendor gave us, else skip the stub row.
                if not part_number:
                    continue
                description = part_number

            qty = _cell_to_qty(ws_val.cell(row=row, column=COL_QTY).value)
            unit_cents = estimate_parse.to_cents(
                ws_val.cell(row=row, column=COL_UNIT_PRICE).value
            )
            if unit_cents is not None and not (0 <= unit_cents <= MAX_CENTS):
                unit_cents = None
            extended_cents = estimate_parse.to_cents(
                ws_val.cell(row=row, column=COL_EXTENDED).value
            )
            if extended_cents is not None and not (0 <= extended_cents <= MAX_CENTS):
                extended_cents = None

            line_math_ok = 1
            if qty is not None and unit_cents is not None:
                computed = _round_half_up(qty * unit_cents)
                if extended_cents is None:
                    extended_cents = computed
                elif extended_cents != computed:
                    line_math_ok = 0
                    all_math_ok = False
            elif extended_cents is None:
                # Unpriced line — carried forward for the disposition screen,
                # no math to check.
                pass
            if extended_cents is not None:
                any_priced = True
                subtotal_cents += extended_cents

            lines.append({
                "position": len(lines) + 1,
                "section": None,
                "part_number": part_number or None,
                "description": description,
                "qty": qty,
                "unit": unit or None,
                "unit_cost_cents": unit_cents,
                "extended_cents": extended_cents,
                "math_ok": line_math_ok,
                "line_note": None,
            })

        meta = _read_meta(wb_val)
        rfq_number = meta.get("ITS_RFQ_NUMBER")
        vendor_key = meta.get("ITS_VENDOR_KEY")
        token = meta.get("ITS_FORM_TOKEN")
        verified = bool(
            rfq_number and vendor_key
            and portal_hmac.verify_rfq_form_token(
                secret, token, rfq_number=rfq_number, vendor_key=vendor_key
            )
        )
        return FormParseResult(
            verified=verified,
            rfq_number=rfq_number if verified else None,
            vendor_key=vendor_key if verified else None,
            lines=lines,
            math_ok=1 if all_math_ok else 0,
            subtotal_cents=subtotal_cents if any_priced else None,
        )
    except Exception:  # noqa: BLE001 — hostile workbook internals; degrade, never raise
        return None
    finally:
        wb_val.close()
        wb_raw.close()
