"""RED-suite tests for po_materials/quote_form.py — the Tier-0 fillable RFQ quote
form (ADR-0004 decision 10, E6/PR-B).

Every test is a PROVE-THE-CONTROL-BITES test:
  * render→fill→parse ROUND TRIP (the golden path): a vendor's programmatic fill
    parses back verified with exact integer-cents math.
  * deterministic render — byte-identical output for fixed inputs (§47 filing).
  * tampered/absent token → verified=False, rfq/vendor identity WITHHELD (the
    auto-bind defense); lines still parse (ordinary ladder upload).
  * red-team #3: a FORMULA in a numeric cell ('=HYPERLINK(...)') rejects the
    WHOLE form (None); formula-lead text in a TEXT field is STRIPPED on
    carry-forward ('=1+1' → '1+1').
  * a zip-bomb xlsx is REFUSED by the §34 screen (the layer that runs BEFORE
    openpyxl in the daemon — the ordering itself is proven end-to-end in
    tests/test_estimate_ladder_wiring.py).
  * the WIDENED _scan_openxml external-relationship gate bites on an
    `externalLink` rel that the OLD (attachedTemplate|oleObject-only) gate
    missed — with the no-rel control fixture staying clean.

Run with: pytest -q tests/test_quote_form.py
"""
from __future__ import annotations

import io
import zipfile

from openpyxl import load_workbook

from po_materials import po_attach_screen, quote_form
from shared import portal_hmac

SECRET = b"quote-form-test-secret"

RFQ = "RFQ-2026-0007"
VENDOR = "PLATT"
JOB = "Sunrise Solar"

LINES = [
    {"part_number": "PVC-200", "description": '2" PVC conduit', "qty": 2, "unit": "EA"},
    {"part_number": "WIRE-THHN", "description": "#12 THHN wire", "qty": 5, "unit": "M"},
]


def _render(
    *,
    rfq_number: str = RFQ,
    vendor_key: str = VENDOR,
    job_name: str = JOB,
    lines: list[dict] = LINES,
    secret: bytes = SECRET,
    due_date: str | None = "2026-08-01",
) -> bytes:
    return quote_form.render_quote_form(
        rfq_number, vendor_key, job_name, lines, secret=secret, due_date=due_date
    )


def _fill(data: bytes, cells: dict[tuple[int, int], object], *, sheet: str = "Quote Form") -> bytes:
    """Programmatically fill cells (the vendor's edit), preserving everything else."""
    wb = load_workbook(io.BytesIO(data))
    ws = wb[sheet]
    for (row, col), value in cells.items():
        ws.cell(row=row, column=col, value=value)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---- determinism + round trip -------------------------------------------------------


def test_render_is_byte_deterministic():
    assert _render() == _render()


def test_render_passes_the_section34_screen():
    """Our own artifact must survive the widened screener (no external rels)."""
    r = po_attach_screen.screen_attachment(
        "quote_form.xlsx", po_attach_screen.MIME_XLSX, _render()
    )
    assert r.disposition == "clean", r


def test_round_trip_golden_fill_parses_verified_with_exact_cents():
    """Render → vendor fills prices → parse: verified identity + integer-cents
    math. Line 1: 2 × $12.50 = $25.00 (computed); line 2: vendor supplies a
    matching extended → math_ok stays 1."""
    filled = _fill(_render(), {
        (9, quote_form.COL_UNIT_PRICE): 12.5,
        (10, quote_form.COL_UNIT_PRICE): "1,098.90",  # text money still parses
        (10, quote_form.COL_EXTENDED): "5,494.50",     # 5 × 1098.90
    })
    parsed = quote_form.parse_quote_form(filled, secret=SECRET)
    assert parsed is not None
    assert parsed.verified is True
    assert (parsed.rfq_number, parsed.vendor_key) == (RFQ, VENDOR)
    assert parsed.math_ok == 1
    [l1, l2] = parsed.lines
    assert (l1["unit_cost_cents"], l1["extended_cents"], l1["math_ok"]) == (1250, 2500, 1)
    assert (l2["unit_cost_cents"], l2["extended_cents"], l2["math_ok"]) == (109890, 549450, 1)
    assert l1["description"] == '2" PVC conduit'
    assert l1["part_number"] == "PVC-200"
    assert parsed.subtotal_cents == 2500 + 549450


def test_mismatched_extended_flags_line_math():
    filled = _fill(_render(), {
        (9, quote_form.COL_UNIT_PRICE): 12.5,
        (9, quote_form.COL_EXTENDED): 99.99,  # 2 × $12.50 is NOT $99.99
    })
    parsed = quote_form.parse_quote_form(filled, secret=SECRET)
    assert parsed is not None
    assert parsed.lines[0]["math_ok"] == 0
    assert parsed.math_ok == 0


# ---- token identity (the auto-bind defense) ----------------------------------------


def test_tampered_meta_yields_verified_false_and_no_identity():
    """Change the RFQ number in _ITS_META: the token no longer matches → the
    parse KEEPS the lines but withholds rfq/vendor (no auto-bind)."""
    tampered = _fill(
        _render(), {(1, 2): "RFQ-9999-EVIL"}, sheet="_ITS_META",
    )
    parsed = quote_form.parse_quote_form(tampered, secret=SECRET)
    assert parsed is not None
    assert parsed.verified is False
    assert parsed.rfq_number is None
    assert parsed.vendor_key is None
    assert len(parsed.lines) == 2  # ordinary ladder upload, lines intact


def test_wrong_secret_yields_verified_false():
    parsed = quote_form.parse_quote_form(_render(), secret=b"some-other-secret")
    assert parsed is not None
    assert parsed.verified is False


def test_token_verify_is_bound_to_both_identity_fields():
    """The rfq-form:v1 token binds (rfq_number, vendor_key): a valid token for
    one pair never verifies another (the replay defense)."""
    token = portal_hmac.rfq_form_token(SECRET, RFQ, VENDOR)
    assert portal_hmac.verify_rfq_form_token(
        SECRET, token, rfq_number=RFQ, vendor_key=VENDOR
    )
    assert not portal_hmac.verify_rfq_form_token(
        SECRET, token, rfq_number=RFQ, vendor_key="NASSAU"
    )
    assert not portal_hmac.verify_rfq_form_token(
        SECRET, token, rfq_number="RFQ-2026-0008", vendor_key=VENDOR
    )
    assert not portal_hmac.verify_rfq_form_token(
        SECRET, None, rfq_number=RFQ, vendor_key=VENDOR
    )


# ---- red-team #3: formula hardening -------------------------------------------------


def test_formula_in_price_cell_rejects_whole_form():
    """'=HYPERLINK(...)' in a numeric input cell → the WHOLE form falls to None
    (never parse a formula as a number). Delete the raw-cell rejection and this
    test fails."""
    filled = _fill(_render(), {
        (9, quote_form.COL_UNIT_PRICE): '=HYPERLINK("http://evil.example","12.50")',
    })
    assert quote_form.parse_quote_form(filled, secret=SECRET) is None


def test_formula_lead_strings_in_numeric_cells_reject_whole_form():
    for hostile in ("=1+1", "+12.50", "-12.50", "@SUM(A1)", "\t12", "\r12"):
        filled = _fill(_render(), {(9, quote_form.COL_QTY): hostile})
        assert quote_form.parse_quote_form(filled, secret=SECRET) is None, hostile


def test_formula_in_description_is_stripped_on_carry_forward():
    """'=1+1' in a TEXT field is not a numeric-cell rejection — the lead chars
    are STRIPPED and the remainder carries forward (CSV/xlsx-injection
    neutralization)."""
    filled = _fill(_render(), {
        (9, quote_form.COL_DESCRIPTION): "=1+1",
        (9, quote_form.COL_UNIT_PRICE): 12.5,
    })
    parsed = quote_form.parse_quote_form(filled, secret=SECRET)
    assert parsed is not None
    assert parsed.lines[0]["description"] == "1+1"


# ---- §34 screen interplay ----------------------------------------------------------


def _zip_bomb_xlsx() -> bytes:
    """An entry-count zip bomb dressed as an xlsx (with the _ITS_META marker, so a
    wrong-ordered pipeline WOULD try to openpyxl-parse it)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "<Types/>")
        zf.writestr("xl/workbook.xml", "<workbook><sheet name='_ITS_META'/></workbook>")
        for i in range(po_attach_screen.MAX_ZIP_ENTRIES + 1):
            zf.writestr(f"xl/junk{i}.xml", "<x/>")
    return buf.getvalue()


def test_zip_bomb_xlsx_refused_by_screen():
    """The §34 screen (which the daemon runs BEFORE any openpyxl parse — ordering
    proven in tests/test_estimate_ladder_wiring.py) refuses the bomb as
    MALICIOUS."""
    r = po_attach_screen.screen_attachment(
        "bomb.xlsx", po_attach_screen.MIME_XLSX, _zip_bomb_xlsx()
    )
    assert r.disposition == "malicious"
    assert r.detail.startswith("zip_entry_bomb")


# ---- the widened external-relationship gate (red-team #3 hardening) -----------------


def _xlsx_with_rels(rels_xml: bytes | None) -> bytes:
    """A structurally-plausible xlsx; optionally carrying a workbook rels part."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "<Types/>")
        zf.writestr("xl/workbook.xml", "<workbook/>")
        if rels_xml is not None:
            zf.writestr("xl/_rels/workbook.xml.rels", rels_xml)
    return buf.getvalue()


_EXTERNAL_LINK_RELS = (
    b'<?xml version="1.0"?><Relationships>'
    b'<Relationship Id="rId9" Type="http://schemas.openxmlformats.org/'
    b'officeDocument/2006/relationships/externalLink" '
    b'Target="externalLinks/externalLink1.xml"/>'
    b"</Relationships>"
)


def test_external_link_rel_is_suspicious_via_the_new_gate():
    """PROVE-THE-NEW-GATE-BITES: an `externalLink` relationship (the xlsx
    external-workbook reference vector) carries NEITHER TargetMode=\"External\"
    NOR the attachedTemplate/oleObject keywords — the OLD gate passed it; the
    widened gate refuses it as suspicious."""
    # The old gate's conditions are absent from the fixture (the RED premise).
    assert b'TargetMode="External"' not in _EXTERNAL_LINK_RELS
    assert b"attachedTemplate" not in _EXTERNAL_LINK_RELS
    assert b"oleObject" not in _EXTERNAL_LINK_RELS
    r = po_attach_screen.screen_attachment(
        "q.xlsx", po_attach_screen.MIME_XLSX, _xlsx_with_rels(_EXTERNAL_LINK_RELS)
    )
    assert (r.disposition, r.detail) == ("suspicious", "openxml_external_relationship")
    assert po_attach_screen.is_structural_active_content(r)


def test_external_reference_and_active_content_external_flag():
    """externalReference still flags; a TargetMode=External relationship flags
    ONLY when it is active-content (attachedTemplate/oleObject), per review F2 —
    the ADR-authorized scope, not every External target."""
    for marker in (
        b'<Relationship Id="r1" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/attachedTemplate" '
        b'Target="evil.dotx" TargetMode="External"/>',
        b'<externalReference r:id="rId1"/>',
    ):
        rels = b"<Relationships>" + marker + b"</Relationships>"
        r = po_attach_screen.screen_attachment(
            "q.xlsx", po_attach_screen.MIME_XLSX, _xlsx_with_rels(rels)
        )
        assert (r.disposition, r.detail) == (
            "suspicious", "openxml_external_relationship"
        ), marker


def test_bare_hyperlink_external_stays_clean():
    """Review F2: an ordinary hyperlink (website/email link) is TargetMode=External
    too. This screener is SHARED with the live PO attachment pass, so flagging every
    hyperlink-bearing vendor doc would flood the Review Queue for no security gain —
    a bare External hyperlink, with no externalLink/externalReference/attachedTemplate/
    oleObject, must stay clean."""
    rels = (
        b'<?xml version="1.0"?><Relationships>'
        b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/hyperlink" '
        b'Target="https://vendor.example.com" TargetMode="External"/>'
        b"</Relationships>"
    )
    r = po_attach_screen.screen_attachment(
        "q.xlsx", po_attach_screen.MIME_XLSX, _xlsx_with_rels(rels)
    )
    assert r.disposition == "clean", r


def test_internal_rels_control_fixture_stays_clean():
    """The control: the SAME container without external markers is clean — the
    new gate flags the marker, not the rels part itself."""
    rels = (
        b'<?xml version="1.0"?><Relationships>'
        b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        b"</Relationships>"
    )
    r = po_attach_screen.screen_attachment(
        "q.xlsx", po_attach_screen.MIME_XLSX, _xlsx_with_rels(rels)
    )
    assert r.disposition == "clean", r
