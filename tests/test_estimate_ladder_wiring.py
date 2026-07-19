"""RED-suite tests for the PR-B extraction-ladder wiring in
po_materials/estimate_poll.py (ADR-0004 E4-E6).

Fully mocked at the module seams (the tests/test_estimate_poll.py house idiom);
the sibling extraction-core modules (estimate_parse / estimate_ocr /
estimate_extract) are patched at their real functions. PROVE-THE-CONTROL-BITES
coverage:

  * ALL tier gates false → PR-A behavior byte-identical: result 'needs_review',
    NO extraction kwarg, ZERO extraction-core calls (the dark-ship contract).
  * tier1 on → 'extracted' posted with the Worker-shape extraction payload
    (INTEGER tier, math_ok 0|1 integers, lines) + Estimate_Log stamped
    'extracted' with the body-derived vendor/quote identity.
  * tier1 math-flagged (needs_review) result → degraded to 'needs_review',
    never posted.
  * tier2 one-doc-per-cycle: two scanned docs, ONE estimate_extract.extract
    call; the second doc lands needs_review.
  * Tier-0 filled-form path END-TO-END REAL (quote_form render + programmatic
    fill + real rfq-form:v1 HMAC): 'extracted' tier 0, payload_json carries the
    verified rfq_number/vendor_key, detail note `rfq_form:<rfq>:<vendor>`
    (the PR-D auto-bind TODO's interim transport).
  * ordering proof (red-team #3): a zip-bomb xlsx is refused by the REAL §34
    screen BEFORE quote_form/openpyxl ever runs (parse mock never called).

Run with: pytest -q tests/test_estimate_ladder_wiring.py
"""
from __future__ import annotations

import base64
import dataclasses
import hashlib
import hmac as _hmac
import json
from types import SimpleNamespace
from typing import Any

import pytest

from po_materials import estimate_poll, po_attach_screen, quote_form
from po_materials.estimate_parse import ExtractionResult, LineItem, ParsedPdf

SECRET = "est-test-secret"

_MINIMAL_PDF = b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog >>\nendobj\ntrailer\n%%EOF\n"

CLEAN = po_attach_screen.ScreenResult("clean", "L2", "ok")

# Captured BEFORE any fixture patches — the real screener for the ordering proof.
REAL_SCREEN = po_attach_screen.screen_attachment

DARK_TIERS = estimate_poll._TierConfig(
    tier1_enabled=False, tier2_enabled=False, ocr_enabled=False,
    model="qwen3.5:9b", ollama_base_url="http://127.0.0.1:11434",
    confidence_threshold=0.75, timeout_seconds=600,
)


def _tiers(**over: Any) -> estimate_poll._TierConfig:
    return dataclasses.replace(DARK_TIERS, **over)


def _sign_est(secret: str, row: dict[str, Any]) -> str:
    canonical = "\n".join([
        "est:v1", row["est_uuid"], row["job_no"], row["filename"],
        row["declared_mime"], str(row["size_bytes"]), row["sha256"],
    ])
    return _hmac.new(secret.encode(), canonical.encode(), hashlib.sha256).hexdigest()


def _est_row(data: bytes, **over: Any) -> dict[str, Any]:
    est_id = int(over.pop("id", 41))
    row: dict[str, Any] = {
        "id": est_id,
        "est_uuid": f"u-est-{est_id}",
        "job_no": str(over.pop("job_no", "2026.001")),
        "job_name": "Sunrise Solar",
        "filename": str(over.pop("filename", "Platt Quote 4471.pdf")),
        "declared_mime": str(over.pop("declared_mime", "application/pdf")),
        "size_bytes": len(data),
        "sha256": hashlib.sha256(data).hexdigest(),
        "status": "pending",
        "uploaded_by": "office.admin",
        "created_at": 1234,
    }
    row["hmac"] = _sign_est(SECRET, row)
    row.update(over)
    return row


def _one_chunk(data: bytes) -> list[dict[str, Any]]:
    return [{"chunk_index": 0, "chunk_total": 1,
             "chunk_b64": base64.b64encode(data).decode()}]


def _good_result(**over: Any) -> ExtractionResult:
    base = ExtractionResult(
        doc_type="quote",
        confidence=0.95,
        line_items=[
            LineItem(description='2" PVC conduit', part_number="PVC-200",
                     qty=2.0, unit="EA", unit_cost_cents=1250,
                     extended_cents=2500, math_ok=True),
        ],
        vendor_name="Platt Electric Supply",
        quote_number="4471",
        subtotal_cents=2500,
        grand_total_cents=2500,
        tier="tier1_generic",
        math_ok=True,
        needs_review=False,
    )
    for key, value in over.items():
        setattr(base, key, value)
    return base


def _parsed_pdf(pages: list[str], *, is_scanned: bool = False) -> ParsedPdf:
    return ParsedPdf(
        pages_text=pages, words=[[] for _ in pages], tables=[[] for _ in pages],
        chars_per_page=[len(p) for p in pages], is_scanned=is_scanned,
    )


@pytest.fixture
def _patch(mocker):
    est_log = mocker.patch("po_materials.estimate_poll.estimate_log")
    est_log.find_row_by_uuid.return_value = None
    est_log.append_row.return_value = 1
    est_log.update_status.return_value = True
    est_log.STATUS_RECEIVED = "received"
    est_log.STATUS_REFUSED = "refused"
    est_log.STATUS_NEEDS_REVIEW = "needs_review"
    est_log.STATUS_EXTRACTED = "extracted"

    upload = mocker.patch(
        "po_materials.estimate_poll.box_client.upload_bytes_or_new_version",
        return_value={"id": "f-est-1", "name": "x", "size": 9},
    )

    seams = {
        "gate": mocker.patch(
            "po_materials.estimate_poll._polling_enabled", return_value=True
        ),
        "resolve_cfg": mocker.patch(
            "po_materials.estimate_poll.resolve_and_log", return_value={}
        ),
        "creds": mocker.patch(
            "po_materials.estimate_poll._resolve_credentials",
            return_value=SimpleNamespace(
                base_url="https://portal.example", bearer="tok", secret=SECRET
            ),
        ),
        "clamav": mocker.patch(
            "po_materials.estimate_poll._attach_clamav_enabled", return_value=False
        ),
        "max_pages": mocker.patch(
            "po_materials.estimate_poll._max_pages_preview", return_value=12
        ),
        # PR-B: the per-cycle tier-config snapshot — tests override return_value.
        "tiers": mocker.patch(
            "po_materials.estimate_poll._resolve_tier_config",
            return_value=DARK_TIERS,
        ),
        "pending": mocker.patch(
            "po_materials.estimate_poll.portal_client.get_estimates_pending",
            return_value=[],
        ),
        "claim": mocker.patch(
            "po_materials.estimate_poll.portal_client.claim_estimate",
            return_value=True,
        ),
        "chunks": mocker.patch(
            "po_materials.estimate_poll.portal_client.get_estimate_chunks",
            return_value=[],
        ),
        "result": mocker.patch(
            "po_materials.estimate_poll.portal_client.post_estimate_result",
            return_value=True,
        ),
        "preview_post": mocker.patch(
            "po_materials.estimate_poll.portal_client.post_estimate_preview",
            return_value=True,
        ),
        "screen": mocker.patch(
            "po_materials.estimate_poll.po_attach_screen.screen_attachment",
            return_value=CLEAN,
        ),
        "extract_text": mocker.patch(
            "po_materials.estimate_classify.extract_pages_text",
            return_value=["QUOTE # 4471\nPlatt Electric Supply\nTotal $25.00"],
        ),
        "classify": mocker.patch(
            "po_materials.estimate_classify.classify_doc_type",
            return_value=("quote", 0.95),
        ),
        "render": mocker.patch(
            "po_materials.estimate_preview.render_page_pngs",
            return_value=[],
        ),
        # Sibling extraction-core seams (the real modules landed in this PR pair).
        "parse_native": mocker.patch(
            "po_materials.estimate_parse.parse_native", return_value=None
        ),
        "templates": mocker.patch(
            "po_materials.estimate_parse.load_vendor_templates", return_value=[]
        ),
        "generic": mocker.patch(
            "po_materials.estimate_parse.parse_generic_table", return_value=None
        ),
        "ocr": mocker.patch(
            "po_materials.estimate_ocr.ocr_pages", return_value=[]
        ),
        "llm": mocker.patch(
            "po_materials.estimate_extract.extract", return_value=None
        ),
        "upload": upload,
        "box_folder": mocker.patch(
            "po_materials.estimate_poll._resolve_quotes_box_folder",
            return_value="folder-est",
        ),
        "est_log": est_log,
        "review_q": mocker.patch(
            "po_materials.estimate_poll.review_queue.add", return_value=1
        ),
        "anomaly": mocker.patch(
            "po_materials.estimate_poll.anomaly_logger.check", return_value=[]
        ),
        "log": mocker.patch("po_materials.estimate_poll.error_log.log", return_value=None),
        "hb": mocker.patch(
            "po_materials.estimate_poll._write_heartbeat", return_value=None
        ),
        "hb_row": mocker.patch(
            "po_materials.estimate_poll._write_heartbeat_row", return_value=None
        ),
        "marker": mocker.patch(
            "po_materials.estimate_poll._write_watchdog_marker", return_value=None
        ),
        "flags_load": mocker.patch(
            "po_materials.estimate_poll._load_flags", return_value={}
        ),
        "flags_persist": mocker.patch(
            "po_materials.estimate_poll._persist_flags", return_value=None
        ),
        "circuit": mocker.patch(
            "po_materials.estimate_poll.circuit_breaker.is_open", return_value=False
        ),
    }
    return seams


def _run(_patch) -> Any:
    return estimate_poll._poll_inside_lock()


def _result_kwargs(_patch, index: int = -1) -> dict[str, Any]:
    assert _patch["result"].call_args_list, "expected a result post"
    return dict(_patch["result"].call_args_list[index].kwargs)


def _est_log_values(_patch) -> str:
    out: list[str] = []
    for call in _patch["est_log"].mock_calls:
        out.extend(str(a) for a in call.args)
        out.extend(f"{k}={v}" for k, v in call.kwargs.items())
    return " | ".join(out)


# ---- 1. dark tiers = PR-A behavior --------------------------------------------------


def test_all_tier_gates_false_is_pr_a_behavior_byte_identical(_patch):
    """The dark-ship contract: with every tier gate false a clean native-text
    quote lands EXACTLY as PR-A shipped it — result 'needs_review', NO
    extraction kwarg, and ZERO extraction-core calls."""
    _patch["pending"].return_value = [_est_row(_MINIMAL_PDF)]
    _patch["chunks"].return_value = _one_chunk(_MINIMAL_PDF)

    stats = _run(_patch)

    result = _result_kwargs(_patch)
    assert result["status"] == "needs_review"
    assert "extraction" not in result
    assert "detail" not in result
    assert result.get("box_file_id") == "f-est-1"
    _patch["parse_native"].assert_not_called()
    _patch["generic"].assert_not_called()
    _patch["ocr"].assert_not_called()
    _patch["llm"].assert_not_called()
    assert stats.extracted == 0
    assert stats.filed == 1
    assert "extracted" not in _est_log_values(_patch)


# ---- 2. tier1 on → extracted posted -------------------------------------------------


def test_tier1_on_posts_extracted_with_worker_shape_payload(_patch):
    _patch["pending"].return_value = [_est_row(_MINIMAL_PDF)]
    _patch["chunks"].return_value = _one_chunk(_MINIMAL_PDF)
    _patch["tiers"].return_value = _tiers(tier1_enabled=True)
    _patch["parse_native"].return_value = _parsed_pdf(["QUOTE # 4471 ..."])
    _patch["generic"].return_value = _good_result()

    stats = _run(_patch)

    result = _result_kwargs(_patch)
    assert result["status"] == "extracted"
    extraction = result["extraction"]
    assert extraction["tier"] == 1  # INTEGER tier (Worker parseExtraction)
    assert extraction["math_ok"] == 1 and extraction["math_ok"] is not True
    assert extraction["doc_type"] == "quote"
    assert extraction["vendor_name"] == "Platt Electric Supply"
    assert extraction["quote_number"] == "4471"
    [line] = extraction["lines"]
    assert (line["position"], line["math_ok"]) == (1, 1)
    assert (line["unit_cost_cents"], line["extended_cents"]) == (1250, 2500)
    assert json.loads(extraction["payload_json"])  # a real JSON document
    assert stats.extracted == 1
    # Estimate_Log stamped extracted with the body-derived identity.
    log_blob = _est_log_values(_patch)
    assert "extracted" in log_blob
    assert "Platt Electric Supply" in log_blob
    assert "4471" in log_blob
    _patch["llm"].assert_not_called()  # tier1 hit never burns the tier2 budget


def test_tier1_math_flagged_result_degrades_to_needs_review(_patch):
    """A Tier-1 result carrying needs_review (math flags) must NOT post as
    extracted — the human disposition screen gets the raw document instead."""
    _patch["pending"].return_value = [_est_row(_MINIMAL_PDF)]
    _patch["chunks"].return_value = _one_chunk(_MINIMAL_PDF)
    _patch["tiers"].return_value = _tiers(tier1_enabled=True)
    _patch["parse_native"].return_value = _parsed_pdf(["QUOTE ..."])
    _patch["generic"].return_value = _good_result(
        math_ok=False, needs_review=True, math_flags=["line 1: mismatch"],
    )

    _run(_patch)

    result = _result_kwargs(_patch)
    assert result["status"] == "needs_review"
    assert "extraction" not in result


# ---- 3. tier2 one-doc-per-cycle -----------------------------------------------------


def test_tier2_budget_is_one_document_per_cycle(_patch):
    """Two SCANNED docs, tier2+ocr on: estimate_extract.extract runs ONCE; the
    second document lands needs_review (the budget, not a failure)."""
    row1 = _est_row(_MINIMAL_PDF, id=41)
    row2 = _est_row(_MINIMAL_PDF, id=42)
    _patch["pending"].return_value = [row1, row2]
    _patch["chunks"].return_value = _one_chunk(_MINIMAL_PDF)
    _patch["tiers"].return_value = _tiers(tier2_enabled=True, ocr_enabled=True)
    _patch["extract_text"].return_value = [""]  # no native text → scanned
    _patch["classify"].return_value = ("other", 0.0)
    _patch["ocr"].return_value = ["OCR PAGE TEXT"]
    _patch["llm"].return_value = _good_result(tier="tier2_llm")

    stats = _run(_patch)

    assert _patch["llm"].call_count == 1  # the one-doc-per-cycle bound
    assert _patch["ocr"].call_count == 1
    statuses = [c.kwargs["status"] for c in _patch["result"].call_args_list]
    assert statuses == ["extracted", "needs_review"]
    assert stats.extracted == 1
    assert stats.filed == 2


def test_tier2_scanned_without_ocr_gate_stays_needs_review(_patch):
    _patch["pending"].return_value = [_est_row(_MINIMAL_PDF)]
    _patch["chunks"].return_value = _one_chunk(_MINIMAL_PDF)
    _patch["tiers"].return_value = _tiers(tier2_enabled=True, ocr_enabled=False)
    _patch["extract_text"].return_value = [""]
    _patch["classify"].return_value = ("other", 0.0)

    _run(_patch)

    _patch["ocr"].assert_not_called()
    _patch["llm"].assert_not_called()
    assert _result_kwargs(_patch)["status"] == "needs_review"


def test_tier2_flagged_needs_review_result_is_not_posted(_patch):
    _patch["pending"].return_value = [_est_row(_MINIMAL_PDF)]
    _patch["chunks"].return_value = _one_chunk(_MINIMAL_PDF)
    _patch["tiers"].return_value = _tiers(tier2_enabled=True, ocr_enabled=True)
    _patch["extract_text"].return_value = [""]
    _patch["classify"].return_value = ("other", 0.0)
    _patch["ocr"].return_value = ["OCR PAGE TEXT"]
    _patch["llm"].return_value = _good_result(
        tier="tier2_llm", needs_review=True, confidence=0.4,
    )

    _run(_patch)

    assert _result_kwargs(_patch)["status"] == "needs_review"
    assert "extraction" not in _result_kwargs(_patch)


# ---- 4. Tier-0 filled-form path (REAL render + fill + HMAC) ------------------------


def _filled_form_bytes() -> bytes:
    import io as _io

    from openpyxl import load_workbook

    rendered = quote_form.render_quote_form(
        "RFQ-2026-0007", "PLATT", "Sunrise Solar",
        [{"part_number": "PVC-200", "description": '2" PVC conduit',
          "qty": 2, "unit": "EA"}],
        secret=SECRET.encode("utf-8"),
    )
    wb = load_workbook(_io.BytesIO(rendered))
    wb["Quote Form"].cell(row=9, column=quote_form.COL_UNIT_PRICE, value=12.5)
    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_tier0_filled_form_posts_extracted_with_rfq_binding(_patch):
    """END-TO-END REAL Tier 0: rendered form + programmatic fill + the real
    rfq-form:v1 HMAC (the daemon's creds.secret) → 'extracted' tier 0 whose
    payload_json carries the verified rfq_number/vendor_key and whose detail
    note is the PR-D interim auto-bind transport."""
    data = _filled_form_bytes()
    _patch["pending"].return_value = [
        _est_row(data, filename="Quote Form (PLATT).xlsx",
                 declared_mime=po_attach_screen.MIME_XLSX)
    ]
    _patch["chunks"].return_value = _one_chunk(data)

    stats = _run(_patch)

    result = _result_kwargs(_patch)
    assert result["status"] == "extracted"
    assert result["detail"] == "rfq_form:RFQ-2026-0007:PLATT"
    extraction = result["extraction"]
    assert extraction["tier"] == 0
    assert extraction["doc_type"] == "filled_form"
    assert extraction["math_ok"] == 1
    payload = json.loads(extraction["payload_json"])
    assert payload["rfq_number"] == "RFQ-2026-0007"
    assert payload["vendor_key"] == "PLATT"
    assert payload["form_verified"] is True
    [line] = extraction["lines"]
    assert (line["unit_cost_cents"], line["extended_cents"]) == (1250, 2500)
    assert stats.extracted == 1
    # No PDF stages for an xlsx: no classification text, no previews.
    _patch["preview_post"].assert_not_called()


def test_tier0_tampered_form_lands_needs_review_no_binding(_patch):
    """A tampered _ITS_META (bad token) form parses but must NOT auto-bind —
    ordinary ladder upload → needs_review (tiers dark)."""
    import io as _io

    from openpyxl import load_workbook

    data = _filled_form_bytes()
    wb = load_workbook(_io.BytesIO(data))
    wb["_ITS_META"]["B1"] = "RFQ-9999-EVIL"
    out = _io.BytesIO()
    wb.save(out)
    tampered = out.getvalue()

    _patch["pending"].return_value = [
        _est_row(tampered, filename="Quote Form (PLATT).xlsx",
                 declared_mime=po_attach_screen.MIME_XLSX)
    ]
    _patch["chunks"].return_value = _one_chunk(tampered)

    _run(_patch)

    result = _result_kwargs(_patch)
    assert result["status"] == "needs_review"
    assert "extraction" not in result


# ---- 5. ordering proof: §34 screen BEFORE openpyxl ---------------------------------


def test_zip_bomb_xlsx_refused_by_real_screen_before_quote_form(mocker, _patch):
    """RED ordering proof (red-team #3): a zip-bomb xlsx carrying the _ITS_META
    marker is refused by the REAL §34 screen and quote_form.parse_quote_form is
    NEVER invoked — the screen precedes every parse. Reorder the pipeline and
    this test fails."""
    import io as _io
    import zipfile as _zipfile

    buf = _io.BytesIO()
    with _zipfile.ZipFile(buf, "w", _zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "<Types/>")
        zf.writestr("xl/workbook.xml", "<workbook><sheet name='_ITS_META'/></workbook>")
        for i in range(po_attach_screen.MAX_ZIP_ENTRIES + 1):
            zf.writestr(f"xl/junk{i}.xml", "<x/>")
    bomb = buf.getvalue()

    parse_spy = mocker.patch(
        "po_materials.quote_form.parse_quote_form",
        side_effect=AssertionError("openpyxl parse reached BEFORE the §34 screen"),
    )
    _patch["screen"].side_effect = REAL_SCREEN  # the real screener judges the bomb

    _patch["pending"].return_value = [
        _est_row(bomb, filename="bomb.xlsx",
                 declared_mime=po_attach_screen.MIME_XLSX)
    ]
    _patch["chunks"].return_value = _one_chunk(bomb)

    _run(_patch)

    parse_spy.assert_not_called()
    _patch["upload"].assert_not_called()
    result = _result_kwargs(_patch)
    assert result["status"] == "refused"
    assert "zip_entry_bomb" in str(result.get("detail", ""))
